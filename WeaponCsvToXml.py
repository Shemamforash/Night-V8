from openpyxl import load_workbook
import string

balance_book = load_workbook('Desolation Balance Files.xlsx', data_only=True)
story_book = load_workbook('Beyond The Veil - Story.xlsx', data_only=True)
num2alpha = dict(zip(range(1, 27), string.ascii_lowercase))


class XMLWriter:
    def __init__(self, sheet_name, output_filename, use_story_book=False):
        if use_story_book:
            self.sheet=story_book[sheet_name]
        else:
            self.sheet = balance_book[sheet_name]
        self.output_file = open('Night Unity Files/Assets/Resources/XML/' + output_filename + ".xml", 'w')

    def __del__(self):
        self.output_file.close()


class WeaponImporter(XMLWriter):
    def __init__(self):
        super(WeaponImporter, self).__init__("Weapon Data V5", "WeaponClasses")
        write_tag(self, "Weapons", self.read_weapon_classes)

    def write_weapon_stats(self, row):
        write_single_value(self, "Damage", get_value(self, "D", row))
        write_single_value(self, "FireRate", get_value(self, "E", row))
        write_single_value(self, "ReloadSpeed", get_value(self, "F", row))
        write_single_value(self, "Accuracy", get_value(self, "G", row))
        write_single_value(self, "Recoil", get_value(self, "H", row))
        write_single_value(self, "Capacity", get_value(self, "I", row))
        write_single_value(self, "Range", get_value(self, "J", row))
        write_single_value(self, "FireType", get_value(self, "P", row))
        write_single_value(self, "FireMode", get_value(self, "Q", row))

    def read_weapon_subtypes(self, subtype_row):
        for i in range(0, 3):
            subtype_name = get_value(self, "C", subtype_row + i)
            automatic = get_value(self, "B", subtype_row + i)
            write_tag(self, "Subtype", self.write_weapon_stats, [subtype_row + i], ["name", "automatic"],
                      [subtype_name, automatic])

    def read_weapon_classes(self):
        for row_no in range(1, 5):
            weapon_class = get_value(self, "A", row_no * 3)
            write_tag(self, "Class", self.read_weapon_subtypes, [row_no * 3], ["name"],
                      [weapon_class])


class RecipeImporter(XMLWriter):
    def __init__(self):
        super(RecipeImporter, self).__init__("Recipes", "Recipes")
        write_tag(self, "Recipes", self.read_recipes)

    def read_recipe(self, row):
        write_single_value(self, "Ingredient1Name", get_value(self, "A", row, "None"))
        write_single_value(self, "Ingredient1Quantity", get_value(self, "B", row, "0"))
        write_single_value(self, "Ingredient2Name", get_value(self, "C", row, "None"))
        write_single_value(self, "Ingredient2Quantity", get_value(self, "D", row, "0"))
        write_single_value(self, "ProductName", get_value(self, "E", row, ""))
        write_single_value(self, "ProductQuantity", get_value(self, "F", row, "0"))
        write_single_value(self, "MakeAll", get_value(self, "G", row))
        write_single_value(self, "Audio", get_value(self, "H", row))
        write_single_value(self, "RecipeType", get_value(self, "I", row))
        write_single_value(self, "LevelNo", get_value(self, "J", row))
        write_single_value(self, "Description", get_value(self, "K", row))

    def read_recipes(self):
        for row_no in range(4, 26):
            write_tag(self, "Recipe", self.read_recipe, [row_no])


class TutorialImporter(XMLWriter):
    def __init__(self):
        super(TutorialImporter, self).__init__("Tutorial", "Tutorial")
        write_tag(self, "Tutorial", self.read_tutorials)

    def read_tutorials(self):
        for row in range(2, 36):
            write_tag(self, "TutorialPart", self.read_single_tutorial, [row])

    def read_single_tutorial(self, row):
        write_single_value(self, "SectionNumber", get_value(self, "A", row))
        write_single_value(self, "SectionName", get_value(self, "B", row))
        write_single_value(self, "PartNumber", get_value(self, "C", row))
        write_single_value(self, "AutoUnlock", get_value(self, "D", row))
        write_single_value(self, "Title", get_value(self, "E", row))
        write_single_value(self, "Text", get_value(self, "F", row))


class ResourceImporter(XMLWriter):
    def __init__(self):
        super(ResourceImporter, self).__init__("Resources", "Resources")
        write_tag(self, "Resources", self.read_resources)

    def read_resource(self, row):
        write_single_value(self, "Name", get_value(self, "A", row, ""))
        write_single_value(self, "Type", get_value(self, "B", row, ""))
        write_single_value(self, "Consumable", get_value(self, "C", row, ""))
        write_single_value(self, "Attribute", get_value(self, "D", row, ""))
        write_single_value(self, "Modifier", get_value(self, "E", row, ""))
        write_single_value(self, "Permanent", get_value(self, "F", row, ""))
        write_single_value(self, "DesertDropRate", get_value(self, "G", row, ""))
        write_single_value(self, "MountainsDropRate", get_value(self, "H", row, ""))
        write_single_value(self, "SeaDropRate", get_value(self, "I", row, ""))
        write_single_value(self, "RuinsDropRate", get_value(self, "J", row, ""))
        write_single_value(self, "WastelandDropRate", get_value(self, "K", row, ""))
        write_single_value(self, "Description", get_value(self, "L", row, ""))
        write_single_value(self, "Effect", get_value(self, "M", row, ""))

    def read_resources(self):
        for row_no in range(3, 37):
            write_tag(self, "Resource", self.read_resource, [row_no])


class InscriptionImporter(XMLWriter):
    def __init__(self):
        super(InscriptionImporter, self).__init__("Inscriptions", "Inscriptions")
        write_tag(self, "Inscriptions", self.read_inscriptions)

    def read_inscriptions(self):
        for row in range(2, 12):
            write_tag(self, "Inscription", self.read_inscription, [row])

    def read_inscription(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "Attribute", get_value(self, "B", row))
        write_single_value(self, "Value", get_value(self, "C", row))


class GearImporter(XMLWriter):
    def __init__(self):
        super(GearImporter, self).__init__("Gear", "Gear")
        write_tag(self, "GearList", self.read_gear)

    def read_gear(self):
        for row in range(3, 12):
            write_tag(self, "Gear", self.read_single_gear, [row])

    def read_single_gear(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "Attribute", get_value(self, "C", row))
        write_single_value(self, "Bonus", get_value(self, "D", row))
        write_single_value(self, "Description", get_value(self, "B", row))


class WeatherImporter(XMLWriter):
    def __init__(self):
        super(WeatherImporter, self).__init__("Weather", "Weather")
        write_tag(self, "WeatherTypes", self.read_weather)

    def read_weather(self):
        for row in range(3, 19):
            write_tag(self, "Weather", self.read_single_weather, [row])

    def read_single_weather(self, row):
        write_single_value(self, "Name", get_value(self, "A", row, "0"))
        write_single_value(self, "DisplayName", get_value(self, "B", row, "0"))
        write_single_value(self, "Temperature", get_value(self, "C", row, "0"))
        write_single_value(self, "Visibility", get_value(self, "D", row, "0"))
        write_single_value(self, "Water", get_value(self, "E", row, "0"))
        write_single_value(self, "Fog", get_value(self, "F", row, "0"))
        write_single_value(self, "Ice", get_value(self, "G", row, "0"))
        write_single_value(self, "Duration", get_value(self, "H", row, "0"))
        write_single_value(self, "Thunder", get_value(self, "I", row, "0"))
        write_tag(self, "Particles", self.read_particles, [row])

    def read_particles(self, row):
        write_single_value(self, "Rain", get_value(self, "J", row, "0"))
        write_single_value(self, "Fog", get_value(self, "K", row, "0"))
        write_single_value(self, "Dust", get_value(self, "L", row, "0"))
        write_single_value(self, "Hail", get_value(self, "M", row, "0"))
        write_single_value(self, "Wind", get_value(self, "N", row, "0"))
        write_single_value(self, "Sun", get_value(self, "O", row, "0"))


class EnvironmentImporter(XMLWriter):
    def __init__(self):
        super(EnvironmentImporter, self).__init__("Environments", "Environments")
        write_tag(self, "EnvironmentTypes", self.read_environment)

    def read_environment(self):
        for row in range(3, 8):
            write_tag(self, get_value(self, "A", row), self.read_single_environment, [row])

    def read_single_environment(self, row):
        write_single_value(self, "Temperature", get_value(self, "C", row))
        write_single_value(self, "Temples", get_value(self, "D", row))
        write_single_value(self, "Monuments", get_value(self, "E", row))
        write_single_value(self, "Shrines", get_value(self, "F", row))
        write_single_value(self, "Fountains", get_value(self, "G", row))
        write_single_value(self, "Shelters", get_value(self, "H", row))
        write_single_value(self, "Animals", get_value(self, "I", row))
        write_single_value(self, "Danger", get_value(self, "J", row))
        write_single_value(self, "WaterSources", get_value(self, "L", row))
        write_single_value(self, "FoodSources", get_value(self, "M", row))
        write_single_value(self, "ResourceSources", get_value(self, "N", row))


class RegionImporter(XMLWriter):
    def __init__(self):
        super(RegionImporter, self).__init__("Region Names", "Regions")
        write_tag(self, "Names", self.read_names)

    def read_names(self):
        for column_num in range(2, 10):
            column_letter = num2alpha[column_num]
            region_type = get_value(self, column_letter, 1, "")
            write_tag(self, region_type, self.read_region_names, [column_letter])

    def read_region_names(self, column_letter):
        write_single_value(self, "Generic", self.read_regions(column_letter, 2, 40))
        write_single_value(self, "Desert", self.read_regions(column_letter, 40, 48))
        write_single_value(self, "Mountains", self.read_regions(column_letter, 48, 59))
        write_single_value(self, "Sea", self.read_regions(column_letter, 59, 71))
        write_single_value(self, "Ruins", self.read_regions(column_letter, 71, 86))
        write_single_value(self, "Wasteland", self.read_regions(column_letter, 86, 104))

    def read_regions(self, column_letter, row_from, row_to):
        names = []
        for row in range(row_from, row_to):
            region_name = get_value(self, column_letter, row, "")
            if region_name == "":
                continue
            names.append(region_name)
        name_string = ","
        name_string = name_string.join(names)
        return name_string


class CharacterImporter(XMLWriter):
    def __init__(self):
        super(CharacterImporter, self).__init__("Classes", "Classes")
        write_tag(self, "Classes", self.read_classes)

    def read_classes(self):
        for row in range(3, 9):
            write_tag(self, "Class", self.read_class, [row])

    def read_class(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "Life", get_value(self, "B", row))
        write_single_value(self, "Will", get_value(self, "C", row))


class EnemyImporter(XMLWriter):
    def __init__(self):
        super(EnemyImporter, self).__init__("Enemy Types", "Enemies")
        write_tag(self, "Enemies", self.read_enemies)

    def read_enemies(self):
        for row in range(3, 22):
            write_tag(self, "Enemy", self.read_enemy, [row])

    def read_enemy(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "DisplayName", get_value(self, "B", row))
        write_single_value(self, "Health", get_value(self, "C", row))
        write_single_value(self, "Speed", get_value(self, "E", row))
        write_single_value(self, "Value", get_value(self, "F", row))
        write_single_value(self, "Difficulty", get_value(self, "G", row))
        write_single_value(self, "DropRate", get_value(self, "H", row))
        write_single_value(self, "Drops", get_value(self, "I", row, ""))
        write_single_value(self, "HasWeapon", get_value(self, "J", row))
        write_single_value(self, "HasGear", get_value(self, "K", row))
        write_single_value(self, "Species", get_value(self, "L", row))


class BrandImporter(XMLWriter):
    def __init__(self):
        super(BrandImporter, self).__init__("Unlockables", "Brands")
        write_tag(self, "Brands", self.read_brands)

    def read_brands(self):
        for row in range(3, 16):
            write_tag(self, get_value(self, "A", row), self.read_brand, [row])

    def read_brand(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "Description", get_value(self, "B", row))
        write_single_value(self, "Requirement", get_value(self, "C", row))
        write_single_value(self, "TargetValue", get_value(self, "D", row))
        write_single_value(self, "Effect", get_value(self, "E", row))
        write_single_value(self, "Modifier", get_value(self, "F", row, "0"))
        write_single_value(self, "RequiresSkill", get_value(self, "G", row))
        write_single_value(self, "MinLevel", get_value(self, "H", row))


class WandererImporter(XMLWriter):
    def __init__(self):
        super(WandererImporter, self).__init__("Wanderer", "Wanderer", True)
        write_tag(self, "Wanderer", self.read_story)

    def read_story(self):
        for row in range(2, 95):
            write_tag(self, "StoryPart", self.read_story_part, [row])

    def read_story_part(self, row):
        write_single_value(self, "Title", get_value(self, "A", row))
        write_single_value(self, "Environment", get_value(self, "B", row))
        write_single_value(self, "PartNumber", get_value(self, "C", row))
        write_single_value(self, "Text", get_value(self, "D", row))


class NecromancerImporter(XMLWriter):
    def __init__(self):
        super(NecromancerImporter, self).__init__("Necromancer", "Necromancer", True)
        write_tag(self, "Necromancer", self.read_story)

    def read_story(self):
        for row in range(2, 28):
            write_tag(self, "StoryPart", self.read_story_part, [row])

    def read_story_part(self, row):
        write_single_value(self, "Title", get_value(self, "A", row))
        write_single_value(self, "PartNumber", get_value(self, "B", row))
        write_single_value(self, "Text", get_value(self, "C", row))


class CharacterStoryImporter(XMLWriter):
    def __init__(self):
        super(CharacterStoryImporter, self).__init__("Characters", "Characters", True)
        write_tag(self, "Characters", self.read_story)

    def read_story(self):
        for row in range(2, 25):
            write_tag(self, "StoryPart", self.read_story_part, [row])

    def read_story_part(self, row):
        write_single_value(self, "Title", get_value(self, "A", row))
        write_single_value(self, "Character", get_value(self, "B", row))
        write_single_value(self, "PartNumber", get_value(self, "C", row))   
        write_single_value(self, "Text", get_value(self, "D", row))


class LoreImporter(XMLWriter):
    def __init__(self):
        super(LoreImporter, self).__init__("Lore", "Lore", True)
        write_tag(self, "Lore", self.read_story)

    def read_story(self):
        for row in range(2, 42):
            write_tag(self, "StoryPart", self.read_story_part, [row])

    def read_story_part(self, row):
        write_single_value(self, "Title", get_value(self, "A", row))
        write_single_value(self, "Unique", get_value(self, "B", row))
        write_single_value(self, "Environment", get_value(self, "C", row))
        write_single_value(self, "Text", get_value(self, "D", row))


class SkillImporter(XMLWriter):
    def __init__(self):
        super(SkillImporter, self).__init__("Skills", "Skills")
        write_tag(self, "Skills", self.read_skills)

    def read_skills(self):
        for row in range(2, 22):
            write_tag(self, "Skill", self.read_skill, [row])

    def read_skill(self, row):
        write_single_value(self, "Name", get_value(self, "A", row))
        write_single_value(self, "Duration", get_value(self, "C", row))
        write_single_value(self, "Cost", get_value(self, "D", row))
        write_single_value(self, "Description", get_value(self, "E", row))


class WeatherProbabilityImporter(XMLWriter):
    def __init__(self):
        super(WeatherProbabilityImporter, self).__init__("Weather", "WeatherProbabilities")
        write_tag(self, "Regions", self.read_regions)

    def read_regions(self):
        write_tag(self, "Desert", self.read_weathers, [20, 27])
        write_tag(self, "Mountains", self.read_weathers, [30, 41])
        write_tag(self, "Sea", self.read_weathers, [44, 54])
        write_tag(self, "Ruins", self.read_weathers, [57, 66])
        write_tag(self, "Wasteland", self.read_weathers, [69, 77])

    def read_weathers(self, row_from, row_to):
        difference = row_to - row_from
        types = ""
        for column in range(2, difference + 1):
            types += get_value(self, num2alpha[column], row_from) + ","
        types = types[:-1]
        write_single_value(self, "Types", types)

        for column in range(2, difference + 1):
            probabilities = ""
            for row in range(row_from + 1, row_to):
                probabilities = probabilities + get_value(self, num2alpha[column], row) + ","
            probabilities = probabilities[:-1]
            write_single_value(self, get_value(self, num2alpha[column], row_from), probabilities)


def write_tag(xml_writer, tag_name, nested_method=None, args=None, parameters=[], values=[]):
    tag = "<" + tag_name
    for parameter_value in zip(parameters, values):
        tag += " " + parameter_value[0] + "=\"" + parameter_value[1] + "\" "
    tag += ">"
    xml_writer.output_file.writelines(tag)
    if nested_method is not None:
        if args is not None:
            nested_method(*args)
        else:
            nested_method()
    xml_writer.output_file.writelines("</" + tag_name + ">")


def get_value(xml_writer, column, row, default_value="1"):
    if column.isdigit():
        column = num2alpha[row]
    value = xml_writer.sheet[column + str(row)].value
    if value is None:
        return default_value
    return str(value)


def write_single_value(xml_writer, stat_name, value):
    value = value.replace("’", "'")
    value = value.replace("‘", "'")
    value = value.replace("”", "\"")
    value = value.replace("“", "\"")
    xml_writer.output_file.write("<" + stat_name + ">" + str(value) + "</" + stat_name + ">")


# WeaponImporter()
GearImporter()
# WeatherImporter()
# WeatherProbabilityImporter()
# RegionImporter()
# CharacterImporter()
EnemyImporter()
# RecipeImporter()
# ResourceImporter()
InscriptionImporter()
# SkillImporter()
# EnvironmentImporter()
# BrandImporter()1
# WandererImporter()
# NecromancerImporter()
# CharacterStoryImporter()
# LoreImporter()
# TutorialImporter()
